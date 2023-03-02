from django.shortcuts import render
from django.http import FileResponse, HttpResponse
from io import BytesIO
from .forms import FileUploadForm , ExcelToPDF,PdfToTxt,bgr,PptTOPdf,MergePdf
from .models import UploadedFile
from docx2pdf import convert
from io import BytesIO
from django.http import FileResponse
from django.shortcuts import render
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import xlrd
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
import os
import PyPDF2
import cv2
import numpy as np
import comtypes
import comtypes.client

comtypes.CoInitialize()

# This function is only for word to pdf
def index(request):
    context = {
        'form': FileUploadForm(),
        'excel_to_pdf': ExcelToPDF(),
        'pdf_to_txt': PdfToTxt(),
        'ppt_to_pdf': PptTOPdf(),
        'bgr': bgr(),
        'merge_pdf': MergePdf()
    }
    return render(request, 'upload_form.html',context)


def upload_word(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.save()
            # Perform processing on the uploaded file here
            input_path = uploaded_file.file.path
            output_path = input_path.replace('.docx', '.pdf')
            convert(input_path, output_path)
            # Return the PDF file as a download
            with open(output_path, 'rb') as pdf_file:
                pdf_bytes = pdf_file.read()
            pdf_file = BytesIO(pdf_bytes)
            response = FileResponse(pdf_file, as_attachment=True, filename=uploaded_file.file.name.replace('.docx', '.pdf'))
            return response
    
      

# This is for xls file and Latest excel
def upload_xls(request):
    if request.method == 'POST':
        # Get the uploaded file from the form
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.save()
            # Load the workbook from the uploaded file
            if uploaded_file.file.name.endswith('.xls'):
                workbook = xlrd.open_workbook(file_contents=uploaded_file.read())
                sheets = workbook.sheet_names()
            elif uploaded_file.file.name.endswith('.xlsx'):
                workbook = load_workbook(uploaded_file.file)
                sheets = workbook.sheetnames
            else:
                raise ValueError('Unsupported file type')
            # Create the PDF file
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            style = ParagraphStyle('table', fontSize=10, leading=12)
            table_style = TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke)
            ])
            for sheet_name in sheets:
                # Get the data from the sheet
                if uploaded_file.file.name.endswith('.xls'):
                    worksheet = workbook.sheet_by_name(sheet_name)
                    data = []
                    for row_num in range(worksheet.nrows):
                        row = worksheet.row_values(row_num)
                        data.append(row)
                elif uploaded_file.file.name.endswith('.xlsx'):
                    worksheet = workbook[sheet_name]
                    data = []
                    for row in worksheet.iter_rows(values_only=True):
                        data.append(row)
                else:
                    raise ValueError('Unsupported file type')
                # Create table and apply style
                table_data = []
                for row in data:
                    table_data.append([Paragraph(str(cell), style) for cell in row])
                table = Table(table_data)
                table.setStyle(table_style)
                elements.append(table)
            # Build and save the PDF file to the default storage
            doc.build(elements)
            buffer.seek(0)
            output_file = ContentFile(buffer.getvalue())
            output_path = uploaded_file.file.name.replace('.xls', '.pdf').replace('.xlsx', '.pdf')
            default_storage.save(output_path, output_file)
            # Return the PDF file as a download
            response = FileResponse(buffer, as_attachment=True, filename=output_path)
            return response

## REMOVE BACKGROUND #############################################

def remove_background(image_path):
    img = cv2.imread(image_path)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV+cv2.THRESH_OTSU)
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    mask = np.zeros(img.shape[:2], np.uint8)
    cv2.drawContours(mask, contours, -1, 255, -1)
    result = cv2.bitwise_and(img, img, mask=mask)
    alpha = np.ones(result.shape[:2], dtype=np.uint8)*255
    alpha[mask == 0] = 0
    result = cv2.cvtColor(result, cv2.COLOR_BGR2BGRA)
    result[:, :, 3] = alpha
    filename = os.path.splitext(image_path)[0] + '.png'
    cv2.imwrite(filename, result)
    return result
  



def remove_background_view(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.save()
            # Perform processing on the uploaded file here
            input_path = uploaded_file.file.path
            output_path = input_path.replace('.png', '_no_bg.png')
            result = remove_background(input_path)
            cv2.imwrite(output_path, result)
            # Return the resulting image as a download
            with open(output_path, 'rb') as image_file:
                image_bytes = image_file.read()
            image_file = BytesIO(image_bytes)
            response = FileResponse(image_file, as_attachment=True, filename=uploaded_file.file.name.replace('.png', '_no_bg.png'))
            response['Content-Type'] = 'image/png'
            return response



def convert_pdf_to_images(pdf_file_path):
    with Image(filename=pdf_file_path) as pdf_image:
        with pdf_image.convert('png') as converted:
            return converted.sequence

def upload_pdf(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.save()
            # Perform processing on the uploaded file here
            input_path = uploaded_file.file.path
            output_path = input_path.replace('.pdf', '.txt')
            with open(input_path, 'rb') as pdf_file:
                pdf_reader = PyPDF2.PdfFileReader(pdf_file)
                num_pages = pdf_reader.getNumPages()
                text = ''
                for i in range(num_pages):
                    page = pdf_reader.getPage(i)
                    text += page.extractText()
            with open(output_path, 'w', encoding='utf-8') as text_file:
                text_file.write(text)
            # Return the text file as a download
            with open(output_path, 'rb') as text_file:
                text_bytes = text_file.read()
            text_file = BytesIO(text_bytes)
            response = FileResponse(text_file, as_attachment=True, filename=uploaded_file.file.name.replace('.pdf', '.txt'))
            return response

def read_pdf(input_path):
    with open(input_path, 'rb') as input_file:
        pdf_reader = PyPDF2.PdfReader(input_file)
        num_pages = len(pdf_reader.pages)
        for page in pdf_reader.pages:
            # Do something with each page
            pass

## PPT TO PDF #############################################


def ppt_to_pdf(input_path, output_path):
    powerpoint = comtypes.client.CreateObject("Powerpoint.Application")
    powerpoint.Visible = True
    slides = powerpoint.Presentations.Open(input_path)
    slides.SaveAs(output_path, FileFormat=32)
    slides.Close()
    powerpoint.Quit()

def upload_file(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.save()
            # Perform processing on the uploaded file here
            input_path = uploaded_file.file.path
            output_path = input_path.replace('.ppt', '.pdf')
            ppt_to_pdf(input_path, output_path)
            # Return the PDF file as a download
            with open(output_path, 'rb') as pdf_file:
                pdf_bytes = pdf_file.read()
            pdf_file = BytesIO(pdf_bytes)
            response = FileResponse(pdf_file, as_attachment=True, filename=uploaded_file.file.name.replace('.ppt', '.pdf'))
            return response
        

def merge_pdf(request):
    if request.method == 'POST':
        # Get all uploaded files
        pdf_files = request.FILES.getlist('pdf_files')
        if len(pdf_files) < 2:
            return render(request, 'error.html', {'error': 'Please select at least two PDF files.'})

        # Merge PDF files
        output_pdf = PyPDF2.PdfFileWriter()
        for pdf in pdf_files:
            input_pdf = PyPDF2.PdfFileReader(pdf)
            for page in range(input_pdf.getNumPages()):
                output_pdf.addPage(input_pdf.getPage(page))
        merged_file_path = os.path.join(os.path.dirname(__file__), 'merged.pdf')
        with open(merged_file_path, 'wb') as file:
            output_pdf.write(file)

        # Return the merged PDF file as a download
        with open(merged_file_path, 'rb') as pdf_file:
            pdf_bytes = pdf_file.read()
        pdf_file = BytesIO(pdf_bytes)
        response = FileResponse(pdf_file, as_attachment=True, filename='merged.pdf')
        return response